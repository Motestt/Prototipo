
import pandas as pd
import os
from xlsxwriter import Workbook
from fpdf import FPDF
import pyexcel as pe
import openpyxl
from openpyxl import load_workbook
from fpdf import FPDF
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont
import xlwings as xw


# metodos:
# metodo preencher(o metodo recebe 4 variaveis como parametro e preenche as sheet em seus devidos lugares)
def preencher(empresa, cliente, o_r, prestador):

    # redireciona os dados para a sheet("tec")
    if empresa == "TECNOLOGIA" and cliente == "BRBPAY":
        excel_nome = 'Principal.xlsx'

        planilha = load_workbook(excel_nome)
        sheet_tec = planilha['tec']

        sheet_tec['E6'] = o_r
        sheet_tec['F6'] = prestador
        # salvando pdf...

        caminho = 'C:\Python\Prototipo\script\PDF_Resultados'
        planilha.save(os.path.join(caminho, f"{o_r}.xlsx"))
        planilha.close()
        # CAMINHO E NOME DOS ARQUIVOS .XLSX
        nome_excel1 = 'C:\Python\Prototipo\script\PDF_Resultados'
        nome_excel = (os.path.join(nome_excel1, f"{o_r}.xlsx"))

        # CAMINHO E NOMES DOS ARQUIVO .PDF
        nome_pdf1 = 'C:\Python\Prototipo\script\PDF_Resultados'
        nome_pdf2 = os.path.join(nome_pdf1, f"{o_r}.pdf")
        nome_pdf = nome_pdf2

       # convertendo para pdf
        wb = xw.Book(nome_excel)
        sheet = wb.sheets['tec']
        sheet.to_pdf(nome_pdf)
        wb.close()

    # redireciona os dados para a sheet("ogea")
    elif empresa == "ÓGEA" and cliente == "ACQIO":

        excel_nome = 'Principal.xlsx'
        planilha = load_workbook(excel_nome)
        sheet_ogea = planilha['ogea']

        sheet_ogea['E7'] = o_r
        sheet_ogea['F7'] = prestador
        # salvando xslx...
        caminho = 'C:\Python\Prototipo\script\PDF_Resultados'
        planilha.save(os.path.join(caminho, f"{o_r}.xlsx"))
        planilha.close()

        nome_excel1 = 'C:\Python\Prototipo\script\PDF_Resultados'
        nome_excel = os.path.join(nome_excel1, f"{o_r}.xlsx")
        # Convertendo o arquivo Excel para PDF
        nome_pdf1 = 'C:\Python\Prototipo\script\PDF_Resultados'
        nome_pdf2 = os.path.join(nome_pdf1, f"{o_r}.pdf")
        nome_pdf = nome_pdf2

        # convertendo para pdf

        wb = xw.Book(nome_excel)
        sheet = wb.sheets['ogea']
        sheet.to_pdf(nome_pdf)
        wb.close()

    else:
        print("Empresa ou cliente inválidos")

# sheetsnames sheets = ['Gerar', 'ogea', 'tec']


# Carregando tabela
tabela1 = pd.read_excel('Principal.xlsx', sheet_name='Gerar')

# usando função:
i = 0
for index,  row in tabela1.iterrows():
    # pegando atributos da tabela ,-,
    empresa = row['Empresa']
    cliente = row['Cliente']
    prestador = row['Prestador']
    o_r = row['OR']

    # metodo "prencher"
    i += 1
    preencher(empresa, cliente, o_r, prestador)


print(f"Foram gerados: {i} .pdf e {i} .xlsx!")
